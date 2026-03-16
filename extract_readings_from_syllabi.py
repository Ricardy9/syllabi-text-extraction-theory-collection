"""
extract_readings_from_syllabi.py

Reads every PDF in "Syllabi to Draw From", sends the extracted text to Claude
(via the Anthropic API) to identify and parse bibliographic references, deduplicates
readings that appear in multiple syllabi, and writes the same structured Excel
spreadsheet as the GROBID-based version.

Required packages:
    pip install anthropic python-dotenv pdfplumber pandas openpyxl

API key is read from .env in the same directory as this script.
Course metadata (title, professor, term, year) is read from
metadata_on_syllabi.xlsx if it exists (produced by metadata_on_syllabi.py).
"""

import json
import logging
import re
import os
import time
from datetime import date
from pathlib import Path

import anthropic
import pdfplumber
import requests
import pandas as pd
import openpyxl
from dotenv import load_dotenv
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Environment / configuration
# ---------------------------------------------------------------------------

load_dotenv(Path(__file__).parent / ".env")

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
MODEL             = "claude-sonnet-4-6"

SYLLABI_FOLDER    = Path(__file__).parent / "Syllabi to Draw From"
METADATA_XLSX     = Path(__file__).parent / "metadata_on_syllabi.xlsx"
OUTPUT_XLSX       = Path(__file__).parent / "literature_from_selected_syllabi.xlsx"
JSON_OUTPUT_DIR   = Path(__file__).parent / "extracted_references"

# Token-level Jaccard threshold for deduplication (same as GROBID version).
TITLE_SIM_THRESHOLD = 0.82

# Crossref polite-pool endpoint — free, no key required.
CROSSREF_API = "https://api.crossref.org/works"
# Minimum Jaccard similarity between query title and Crossref result title
# before we trust the DOI match.
CROSSREF_TITLE_THRESHOLD = 0.70

COLUMNS = [
    "Original Citation",
    "Author(s)",
    "Year",
    "Title",
    "Journal / Publisher",
    "Vol. / Issue / Pages",
    "Link",
    "Source Type",
    "Date Added",
    "Class/es where listed on syllabus",
]

logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
log = logging.getLogger(__name__)

class _HttpRequestNumberer(logging.Filter):
    """Rewrites httpx 'HTTP Request' log lines to include a running count."""
    def __init__(self):
        super().__init__()
        self._count = 0

    def filter(self, record):
        if "HTTP Request" in record.getMessage():
            self._count += 1
            record.msg = record.msg.replace(
                "HTTP Request", f"HTTP Request #{self._count}", 1
            )
            record.args = ()
        return True

_numberer = _HttpRequestNumberer()
logging.getLogger("httpx").addFilter(_numberer)

TODAY = date.today().isoformat()

# ---------------------------------------------------------------------------
# PDF text extraction
# ---------------------------------------------------------------------------

def extract_pdf_text(pdf_path: Path) -> str:
    """
    Extract all text from a PDF using pdfplumber.
    Returns an empty string if extraction fails or the PDF is image-only.
    """
    pages = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    pages.append(text)
    except Exception as exc:
        log.error("  PDF extraction failed for %s: %s", pdf_path.name, exc)
        return ""
    full_text = "\n".join(pages)
    if not full_text.strip():
        log.warning(
            "  No text extracted from %s — the PDF may be image-only (scanned). "
            "Claude cannot process it without OCR.",
            pdf_path.name,
        )
    return full_text


# ---------------------------------------------------------------------------
# Claude extraction
# ---------------------------------------------------------------------------

# Each chunk of syllabus text sent in one API call.  At ~4 chars/token a
# 6 000-char chunk uses ≈1 500 input tokens, keeping JSON output (~25–35
# refs) well within the 8 192-token output limit and avoiding truncation.
CHUNK_SIZE = 6_000
# Minimum chunk size when retrying a failed chunk by splitting.
# At ~4 chars/token a 1 500-char chunk is ~375 tokens — small enough that
# even a very dense passage will not exceed the output token limit.
MIN_RETRY_CHUNK = 1_500

SYSTEM_PROMPT = """\
You are a bibliographic reference parser specialising in academic syllabi.
The user will provide a portion of a course syllabus. Your job is to extract
every assigned reading (article, book, book chapter, working paper, report, etc.)
that a student would be expected to read.

For each reading return a JSON object with exactly these keys:
  - "original_citation"  : the full citation text as it appears verbatim in the syllabus
  - "authors"            : author name(s), separated by "; ", each in "Last, First" format
                           (e.g. "Acemoglu, Daron; Verdier, Thierry"). Convert to Last, First
                           regardless of how the name appears in the syllabus.
  - "year"               : four-digit publication year as a string, or "" if unknown
  - "title"              : title of the article, chapter, or book, normalized by language:
                           - English: proper title case (capitalize first and last word, all major
                             words; lowercase articles, coordinating conjunctions, and prepositions
                             of fewer than five letters unless they open the title, e.g.
                             "The Simple Economics of Extortion: Evidence from Trucking in Aceh")
                           - Spanish (and other Romance languages): sentence case (capitalize only
                             the first word of the title and subtitle, and proper nouns,
                             e.g. "La economía política de las instituciones")
  - "journal"            : journal name, book title, or publisher, as appropriate; "" if none
  - "issue_pages"        : volume/issue/page information (e.g. "Vol. 12, No. 3, pp. 45–67"); "" if none
  - "link"               : DOI URL or other URL if present in the citation; "" if none
  - "source_type"        : one of "Journal Article", "Book", "Book Chapter", "Working Paper", "Report", "White Paper"

Return ONLY a valid JSON array of these objects — no markdown fences, no extra text.
If you cannot identify any readings, return an empty array: []

Rules:
- Include only actual assigned readings, not references cited within those readings.
- Exclude section headers, course schedule rows with no citation, instructor contact info, etc.
- Symbols like §, *, ** used to mark required/optional readings should be stripped from all fields.
- If a field is genuinely unknown, use an empty string "".
- Do not invent information that is not in the text, EXCEPT for titles: if you recognise
  a well-known academic work and the syllabus only shows a shortened or subtitle-only
  version, supply the complete canonical title in the "title" field.
- IMPORTANT: JSON string values must never contain unescaped straight double-quote characters (").
  If citation text includes quotation marks around an article title or any other text, replace
  them with curly/typographic quotes (\u201c and \u201d) instead of escaping with backslash.
  Example: "The Unequal Effects" should become \u201cThe Unequal Effects\u201d in the JSON value.
"""


def _chunk_text(text: str) -> list[str]:
    """
    Split syllabus text into chunks of at most CHUNK_SIZE characters,
    always breaking at a newline so citations are never split mid-line.
    """
    if len(text) <= CHUNK_SIZE:
        return [text]
    chunks = []
    while text:
        if len(text) <= CHUNK_SIZE:
            chunks.append(text)
            break
        split_at = text.rfind("\n", 0, CHUNK_SIZE)
        if split_at == -1:
            split_at = CHUNK_SIZE
        chunks.append(text[:split_at])
        text = text[split_at:].lstrip("\n")
    return chunks


def _repair_json_inner_quotes(raw: str) -> str:
    """
    Replace unescaped straight double-quote characters that appear *inside*
    JSON string values with curly/typographic quotes (U+201C / U+201D), so
    that json.loads() can parse the result.

    Uses a state machine that tracks whether the scanner is currently inside
    a JSON string.  When a `"` is encountered inside a string, it uses the
    following lookahead to decide whether it is a structural closing quote or
    an inner content quote:

      After the `"`, skip any whitespace, then:
        - `}` or `]` or `:` or end-of-input  → structural closing quote.
        - `,`  → peek further past the comma + whitespace:
              if the next non-whitespace char is `"` (the start of the next
              JSON key/value), it is a structural separator  → closing quote.
              if it is anything else (e.g. a letter, as in `", AER."`), the
              comma is citation punctuation inside the string  → content quote.
        - anything else  → content quote.

    This handles the common academic-citation pattern of titles wrapped in
    straight double-quotes followed by a comma (e.g. `"Title," Journal`).
    """
    result   = []
    in_str   = False
    open_cq  = True   # alternate between opening (\u201c) and closing (\u201d)
    i        = 0
    n        = len(raw)

    while i < n:
        c = raw[i]

        if in_str:
            if c == "\\":
                # Escape sequence — pass through both characters verbatim.
                result.append(c)
                i += 1
                if i < n:
                    result.append(raw[i])
                    i += 1
                continue
            elif c == '"':
                # Classify: structural closing quote, or inner content quote?
                j = i + 1
                while j < n and raw[j] in " \t\r\n":
                    j += 1
                next_ch = raw[j] if j < n else ""

                if next_ch in "}]:" or j >= n:
                    is_closing = True
                elif next_ch == ",":
                    # Look past the comma + whitespace.
                    k = j + 1
                    while k < n and raw[k] in " \t\r\n":
                        k += 1
                    # Structural separator: next token starts with " (key or value).
                    is_closing = (k < n and raw[k] == '"')
                else:
                    is_closing = False

                if is_closing:
                    in_str  = False
                    open_cq = True
                    result.append('"')
                else:
                    result.append("\u201c" if open_cq else "\u201d")
                    open_cq = not open_cq
            else:
                result.append(c)
        else:
            if c == '"':
                in_str = True
                result.append(c)
            else:
                result.append(c)
        i += 1

    return "".join(result)


def _call_claude_for_chunk(
    chunk: str,
    pdf_name: str,
    chunk_label: str,
    client: anthropic.Anthropic,
) -> tuple[list[dict], bool]:
    """
    Send one text chunk to Claude.
    Returns (refs, had_json_error) — had_json_error=True triggers a retry.
    API errors are not retried (they indicate a connectivity / auth problem).
    """
    try:
        message = client.messages.create(
            model=MODEL,
            max_tokens=8192,
            system=SYSTEM_PROMPT,
            messages=[
                {
                    "role": "user",
                    "content": (
                        f"Extract all assigned readings from the following syllabus excerpt "
                        f"({pdf_name}, {chunk_label}):\n\n{chunk}"
                    ),
                }
            ],
        )
    except anthropic.APIError as exc:
        log.error("  Anthropic API error for %s (%s): %s", pdf_name, chunk_label, exc)
        return [], False  # API error — don't retry by splitting

    raw = message.content[0].text.strip()
    raw = re.sub(r"^```[a-zA-Z]*\n?", "", raw)
    raw = re.sub(r"\n?```$", "", raw)

    # Claude sometimes prepends explanatory text before the JSON array.
    # Extract just the array portion so those responses still parse correctly.
    bracket_start = raw.find("[")
    bracket_end   = raw.rfind("]")
    if bracket_start != -1 and bracket_end > bracket_start:
        raw = raw[bracket_start : bracket_end + 1]

    try:
        refs = json.loads(raw)
    except json.JSONDecodeError as exc:
        # --- Repair attempt: unescaped double-quotes inside string values ----
        # Claude occasionally includes literal " characters from the PDF text
        # (e.g. article titles wrapped in quotes per AER citation style) without
        # escaping them.  The stateful scanner replaces inner content quotes with
        # curly/typographic equivalents so json.loads() can parse the result.
        repaired = _repair_json_inner_quotes(raw)
        try:
            refs = json.loads(repaired)
            log.info(
                "  Repaired inner quotes in %s (%s) — %d ref(s) recovered.",
                pdf_name, chunk_label, len(refs) if isinstance(refs, list) else 0,
            )
        except json.JSONDecodeError as repair_exc:
            log.error(
                "  Invalid JSON from Claude for %s (%s) (%s). First 300 chars:\n%s",
                pdf_name, chunk_label, exc, raw[:300],
            )
            # Trigger retry-by-splitting for any JSON error that survives repair.
            # A chunk with multiple broken citations can't be fixed in one pass;
            # splitting isolates each broken citation into its own sub-chunk where
            # the scanner can handle it individually.
            # Exception: if the chunk is already at minimum size, flag as a leaf
            # failure rather than looping forever (handled by the caller).
            return [], True

    if not isinstance(refs, list):
        return [], False

    cleaned = []
    for item in refs:
        if not isinstance(item, dict):
            continue
        ref = {
            "original_citation": str(item.get("original_citation", "")).strip(),
            "authors":           str(item.get("authors",           "")).strip(),
            "year":              str(item.get("year",              "")).strip(),
            "title":             str(item.get("title",             "")).strip(),
            "journal":           str(item.get("journal",           "")).strip(),
            "issue_pages":       str(item.get("issue_pages",       "")).strip(),
            "link":              str(item.get("link",              "")).strip(),
            "source_type":       str(item.get("source_type",       "")).strip(),
        }
        if ref["title"]:
            cleaned.append(ref)
    return cleaned, False


def _extract_with_retry(
    chunk: str,
    pdf_name: str,
    chunk_label: str,
    client: anthropic.Anthropic,
) -> tuple[list[dict], int]:
    """
    Try to extract refs from a chunk.  On JSON error (truncated output),
    split the chunk in half and retry each half recursively, down to
    MIN_RETRY_CHUNK characters.

    Returns (refs, failed_leaf_count) where failed_leaf_count > 0 means
    at least one atomic sub-chunk could not be parsed even at minimum size.
    """
    refs, had_error = _call_claude_for_chunk(chunk, pdf_name, chunk_label, client)
    if not had_error:
        return refs, 0

    # JSON error — split and retry if the chunk is large enough
    if len(chunk) <= MIN_RETRY_CHUNK:
        log.warning(
            "  ⚠  Chunk '%s' (%d chars) is too small to split further — "
            "some references from this passage may be missing.",
            chunk_label, len(chunk),
        )
        return [], 1  # one irrecoverable failure

    log.info(
        "  JSON error in '%s' (%d chars) — splitting and retrying…",
        chunk_label, len(chunk),
    )
    mid = chunk.rfind("\n", 0, len(chunk) // 2)
    if mid == -1:
        mid = len(chunk) // 2
    half_a = chunk[:mid].strip()
    half_b = chunk[mid:].lstrip("\n").strip()

    refs_a, failed_a = _extract_with_retry(half_a, pdf_name, chunk_label + "a", client)
    refs_b, failed_b = _extract_with_retry(half_b, pdf_name, chunk_label + "b", client)
    return refs_a + refs_b, failed_a + failed_b


def extract_references_with_claude(
    syllabus_text: str,
    pdf_name: str,
    client: anthropic.Anthropic,
) -> tuple[list[dict], dict]:
    """
    Split the syllabus into chunks, send each to Claude (with automatic
    retry-by-splitting on JSON errors), and combine results.

    Returns (refs, stats) where stats = {"chunks": N, "failed": M}.
    failed > 0 means M atomic sub-chunks could not be parsed even after
    recursive splitting — those passages may have missing references.
    """
    if not syllabus_text.strip():
        return [], {"chunks": 0, "failed": 0}

    chunks = _chunk_text(syllabus_text)
    log.info(
        "  %d characters → %d chunk(s) for Claude.",
        len(syllabus_text), len(chunks),
    )

    all_refs: list[dict] = []
    total_failed = 0
    for i, chunk in enumerate(chunks, start=1):
        label = f"part {i}/{len(chunks)}"
        refs, failed = _extract_with_retry(chunk, pdf_name, label, client)
        suffix = f" ({failed} sub-chunk(s) still failed after retry)" if failed else ""
        log.info("  %s → %d reference(s)%s.", label, len(refs), suffix)
        all_refs.extend(refs)
        total_failed += failed

    log.info("  Total extracted from %s: %d reference(s).", pdf_name, len(all_refs))
    return all_refs, {"chunks": len(chunks), "failed": total_failed}


# ---------------------------------------------------------------------------
# Crossref DOI enrichment
# ---------------------------------------------------------------------------

def _crossref_lookup(title: str, authors: str) -> str:
    """
    Query the Crossref API for a DOI matching the given title (and optionally
    first author surname).  Returns a 'https://doi.org/...' string on a
    confident match, or '' on failure / low-confidence.
    """
    # Build a bibliographic query string
    first_author = authors.split(";")[0].split(",")[0].strip() if authors else ""
    query = f"{title} {first_author}".strip()

    try:
        resp = requests.get(
            CROSSREF_API,
            params={"query.bibliographic": query, "rows": 1, "select": "DOI,title"},
            headers={"User-Agent": "syllabus-literature-extractor/1.0 (academic research tool)"},
            timeout=10,
        )
        if resp.status_code != 200:
            return ""
        items = resp.json().get("message", {}).get("items", [])
        if not items:
            return ""
        item = items[0]
        # Verify the returned title is actually a good match
        result_titles = item.get("title", [])
        if not result_titles:
            return ""
        result_title = result_titles[0]
        if _jaccard(_normalise(title), _normalise(result_title)) < CROSSREF_TITLE_THRESHOLD:
            return ""
        doi = item.get("DOI", "")
        return f"https://doi.org/{doi}" if doi else ""
    except Exception:
        return ""


def enrich_links(refs: list[dict]) -> list[dict]:
    """
    For every reference with an empty 'link' field, attempt a Crossref
    lookup.  Modifies the list in place and returns it.
    A short sleep between requests keeps us in Crossref's polite pool.
    """
    missing = [r for r in refs if not r.get("link")]
    if not missing:
        return refs
    log.info("  Looking up DOIs via Crossref for %d reference(s) with no link…", len(missing))
    found = 0
    for ref in missing:
        doi_url = _crossref_lookup(ref["title"], ref.get("authors", ""))
        if doi_url:
            ref["link"] = doi_url
            found += 1
        time.sleep(0.12)   # ~8 req/s — well within Crossref's polite limits
    log.info("  Crossref enrichment: found DOIs for %d/%d reference(s).", found, len(missing))
    return refs


# ---------------------------------------------------------------------------
# Intermediate JSON persistence
# ---------------------------------------------------------------------------

def save_syllabus_json(refs: list[dict], pdf_stem: str) -> None:
    """
    Save the extracted references for one syllabus as a JSON file in
    JSON_OUTPUT_DIR/<pdf_stem>.json.  Creates the directory if needed.
    """
    JSON_OUTPUT_DIR.mkdir(exist_ok=True)
    out_path = JSON_OUTPUT_DIR / f"{pdf_stem}.json"
    with open(out_path, "w", encoding="utf-8") as fh:
        json.dump(refs, fh, ensure_ascii=False, indent=2)
    log.info("  Saved intermediate JSON → %s", out_path.name)


# ---------------------------------------------------------------------------
# Deduplication
# ---------------------------------------------------------------------------

def _normalise(title: str) -> str:
    t = title.lower()
    t = re.sub(r"[^\w\s]", " ", t)
    t = re.sub(r"\b(a|an|the)\b", " ", t)
    return re.sub(r"\s+", " ", t).strip()


def _jaccard(a: str, b: str) -> float:
    sa = set(a.split())
    sb = set(b.split())
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / len(sa | sb)


def merge_references(all_refs: list[tuple[str, dict]]) -> list[dict]:
    """
    Deduplicate references across syllabi using token-level Jaccard similarity
    on normalised titles.  Returns a list of unique ref dicts, each with a
    'classes' key listing every course that assigned the reading.
    """
    merged:     list[dict] = []
    norm_index: list[str]  = []

    for class_label, ref in all_refs:
        norm = _normalise(ref["title"])

        best_idx   = -1
        best_score = 0.0
        for i, existing_norm in enumerate(norm_index):
            score = _jaccard(norm, existing_norm)
            if score > best_score:
                best_score = score
                best_idx   = i

        if best_score >= TITLE_SIM_THRESHOLD and best_idx >= 0:
            classes = merged[best_idx]["classes"]
            if class_label and class_label not in classes:
                classes.append(class_label)
        else:
            entry = dict(ref)
            entry["classes"]    = [class_label] if class_label else []
            entry["date_added"] = TODAY
            merged.append(entry)
            norm_index.append(norm)

    return merged


# ---------------------------------------------------------------------------
# Syllabus metadata loader
# ---------------------------------------------------------------------------

def load_class_labels(metadata_path: Path) -> dict[str, str]:
    if not metadata_path.exists():
        log.warning(
            "Metadata file not found at %s. "
            "Run metadata_on_syllabi.py first for richer class labels. "
            "Falling back to raw PDF filenames.",
            metadata_path,
        )
        return {}

    df = pd.read_excel(metadata_path)
    col_pdf      = "Original name of syllabus PDF"
    col_new_name = "New Syllabus Name"

    labels: dict[str, str] = {}
    for _, row in df.iterrows():
        pdf_name = str(row.get(col_pdf, "")).strip()
        if not pdf_name:
            continue
        new_name = str(row.get(col_new_name, "")).strip()
        labels[pdf_name] = new_name if new_name else pdf_name

    return labels


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
BODY_FONT   = Font(name="Calibri", size=10)
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
WRAP        = Alignment(wrap_text=True, vertical="top")
THIN_SIDE   = Side(style="thin", color="B0C4DE")
CELL_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE
)

COL_WIDTHS = {
    "A": 45,  # Original Citation
    "B": 30,  # Author(s)
    "C": 8,   # Year
    "D": 55,  # Title
    "E": 32,  # Journal / Publisher
    "F": 22,  # Vol. / Issue / Pages
    "G": 38,  # Link
    "H": 18,  # Source Type
    "I": 14,  # Date Added
    "J": 65,  # Class/es
}


def write_excel(
    merged: list[dict],
    all_class_labels: list[str],
    out_path: Path,
) -> None:
    wb = openpyxl.Workbook()

    # Sheet 1: Literature
    ws = wb.active
    ws.title = "Literature"

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = WRAP
        cell.border    = CELL_BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    for row_idx, ref in enumerate(merged, start=2):
        classes_str = "; ".join(ref.get("classes", []))
        values = [
            ref.get("original_citation", ""),
            ref.get("authors",           ""),
            ref.get("year",              ""),
            ref.get("title",             ""),
            ref.get("journal",           ""),
            ref.get("issue_pages",       ""),
            ref.get("link",              ""),
            ref.get("source_type",       ""),
            ref.get("date_added",        TODAY),
            classes_str,
        ]
        row_fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = BODY_FONT
            cell.alignment = WRAP
            cell.border    = CELL_BORDER
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[row_idx].height = 45

    # Sheet 2: _ClassList (for data validation)
    ws_cls = wb.create_sheet("_ClassList")
    ws_cls.cell(row=1, column=1, value="Available Classes").font = Font(
        bold=True, name="Calibri", size=11
    )
    sorted_classes = sorted(all_class_labels)
    for i, label in enumerate(sorted_classes, start=2):
        ws_cls.cell(row=i, column=1, value=label).font = BODY_FONT
    ws_cls.column_dimensions["A"].width = 70

    last_row = max(len(merged) + 1, 2)
    extra    = 500

    # Source Type dropdown (column H)
    dv_type = DataValidation(
        type="list",
        formula1='"Book,Book Chapter,Journal Article,Working Paper,Report,White Paper"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_type.sqref            = f"H2:H{last_row + extra}"
    dv_type.prompt           = "Choose source type"
    dv_type.promptTitle      = "Source Type"
    dv_type.showErrorMessage = False
    ws.add_data_validation(dv_type)

    # Classes dropdown (column J)
    if all_class_labels:
        class_formula = f"_ClassList!$A$2:$A${len(sorted_classes) + 1}"
        dv_class = DataValidation(
            type="list",
            formula1=class_formula,
            allow_blank=True,
            showDropDown=False,
        )
        dv_class.sqref            = f"J2:J{last_row + extra}"
        dv_class.prompt           = (
            "Select a class from the dropdown, or type multiple classes "
            "separated by '; ' for readings that appear in more than one course."
        )
        dv_class.promptTitle      = "Class/es"
        dv_class.showErrorMessage = False
        ws.add_data_validation(dv_class)

    wb.save(out_path)
    log.info("Saved spreadsheet → %s", out_path)


# ---------------------------------------------------------------------------
# Extraction quality report
# ---------------------------------------------------------------------------

def _print_quality_table(syllabus_stats: list[dict]) -> None:
    """
    Print a per-syllabus table showing chunk counts, failures, and refs
    extracted.  Any row with failed > 0 is flagged so the user knows that
    some references from that syllabus may be missing.
    """
    W_LABEL   = 55
    W_CHUNKS  = 8
    W_FAILED  = 8
    W_REFS    = 6
    sep = "=" * (W_LABEL + W_CHUNKS + W_FAILED + W_REFS + 7)

    print("\n" + sep)
    print("  EXTRACTION QUALITY CHECK — per-syllabus chunk failure report")
    print(sep)
    header = (
        f"  {'Syllabus':<{W_LABEL}}  "
        f"{'Chunks':>{W_CHUNKS}}  "
        f"{'Failed':>{W_FAILED}}  "
        f"{'Refs':>{W_REFS}}"
    )
    print(header)
    print("  " + "-" * (W_LABEL + W_CHUNKS + W_FAILED + W_REFS + 6))

    any_failures = False
    for row in syllabus_stats:
        label   = row["label"][:W_LABEL]
        chunks  = row["chunks"]
        failed  = row["failed"]
        refs    = row["refs"]
        flag    = "  ⚠  MISSING REFS POSSIBLE" if failed else ""
        if failed:
            any_failures = True
        print(
            f"  {label:<{W_LABEL}}  "
            f"{chunks:>{W_CHUNKS}}  "
            f"{failed:>{W_FAILED}}  "
            f"{refs:>{W_REFS}}"
            f"{flag}"
        )

    print(sep)
    if any_failures:
        print(
            "  ⚠  One or more syllabi had chunks that could NOT be parsed even\n"
            "     after recursive splitting.  References from those passages are\n"
            "     missing from the spreadsheet.  Check the log output above for\n"
            "     details on which chunk(s) failed."
        )
    else:
        print("  ✓  All chunks parsed successfully — no references dropped.")
    print(sep + "\n")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    if not ANTHROPIC_API_KEY:
        log.error(
            "ANTHROPIC_API_KEY not found. "
            "Make sure it is set in the .env file in the same directory as this script."
        )
        return

    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    pdf_files = sorted(SYLLABI_FOLDER.glob("*.pdf"))
    if not pdf_files:
        log.error("No PDF files found in: %s", SYLLABI_FOLDER)
        return

    log.info("Found %d PDF(s) in '%s'", len(pdf_files), SYLLABI_FOLDER)

    class_labels = load_class_labels(METADATA_XLSX)
    all_class_label_set: set[str] = set()
    all_refs: list[tuple[str, dict]] = []
    syllabus_stats: list[dict] = []

    total_pdfs = len(pdf_files)
    for pdf_idx, pdf_path in enumerate(pdf_files, start=1):
        filename    = pdf_path.name
        class_label = class_labels.get(filename, filename)
        all_class_label_set.add(class_label)

        log.info("Processing: %s  →  class label: %s", filename, class_label)

        syllabus_text = extract_pdf_text(pdf_path)
        if not syllabus_text.strip():
            log.warning("  Skipping %s (no extractable text).", filename)
            syllabus_stats.append({"label": class_label, "chunks": 0, "failed": 0, "refs": 0})
            log.info("  ── %d/%d syllabi parsed ──", pdf_idx, total_pdfs)
            continue

        refs, stats = extract_references_with_claude(syllabus_text, filename, client)
        enrich_links(refs)
        save_syllabus_json(refs, pdf_path.stem)
        for ref in refs:
            all_refs.append((class_label, ref))
        syllabus_stats.append({
            "label":  class_label,
            "chunks": stats["chunks"],
            "failed": stats["failed"],
            "refs":   len(refs),
        })
        log.info("  ── %d/%d syllabi parsed ──", pdf_idx, total_pdfs)

    log.info("Total raw references across all syllabi: %d", len(all_refs))

    merged = merge_references(all_refs)
    log.info("After deduplication: %d unique reference(s)", len(merged))


    write_excel(merged, list(all_class_label_set), OUTPUT_XLSX)

    _print_quality_table(syllabus_stats)

    print("\n" + "=" * 58)
    print("  LITERATURE EXTRACTION SUMMARY  (Claude API)")
    print("=" * 58)
    print(f"  PDFs processed             : {len(pdf_files)}")
    print(f"  Raw references extracted   : {len(all_refs)}")
    print(f"  Unique references (deduped): {len(merged)}")
    print(f"  Output spreadsheet         : {OUTPUT_XLSX.name}")
    print("=" * 58 + "\n")


if __name__ == "__main__":
    main()
